import openpyxl
import random

def readParameters(wb):
    parametersSheet = wb['Parameters']
    envRows = int(parametersSheet.cell(row=1, column=2).value)
    envColumns = int(parametersSheet.cell(row=2, column=2).value)
    envRules = int(parametersSheet.cell(row=3, column=2).value)
    steps = int(parametersSheet.cell(row=4, column=2).value)
    animationDelay = float(parametersSheet.cell(row=5, column=2).value)
    rowIndex = 8
    jokerSymbols = {}
    while parametersSheet.cell(row=rowIndex, column=1).value != 'JokerSymbolsEnd':
        jokerSymbols[str(parametersSheet.cell(row=rowIndex, column=2).value)] = str(parametersSheet.cell(row=rowIndex, column=3).value).split(',')
        rowIndex += 1
    rowIndex += 1
    colorSettings = {}
    while parametersSheet.cell(row=rowIndex, column=1).value != 'ColorSettingsEnd':
        colorSettings[str(parametersSheet.cell(row=rowIndex, column=2).value)] = str(parametersSheet.cell(row=rowIndex, column=3).value).strip()
        rowIndex += 1
    return {
        'envRows': envRows,
        'envColumns': envColumns,
        'envRules': envRules,
        'steps': steps,
        'animationDelay': animationDelay,
        'jokerSymbols': jokerSymbols,
        'colorSettings': colorSettings
    }

def readEnvironment(wb, rows, columns, rules):
    environmentSheet = wb['Environment']
    # make a boundary of the environmant from an empty strings
    envMatrix = [['' for j in range(columns+2)] for i in range(rows+2)]
    for i in range(1, rows + 1):
        for j in range(1, columns + 1):
            envMatrix[i][j] = str(environmentSheet.cell(row=i, column=j).value).split(',')
    envRulesSheet = wb['Environment_rules']
    envRules = []
    for i in range(1, rules + 1):
        rule = {
            'left': str(envRulesSheet.cell(row=i, column=1).value).split(','),
            'right': str(envRulesSheet.cell(row=i, column=3).value).split(',')
        }
        envRules.append(rule)
    return {
        'matrix': envMatrix,
        'rules': envRules
    }

def readAgents(wb, rows, columns):
    agentsSheet  = wb['Agents']
    agents = []
    lineIndex = 1
    while agentsSheet.cell(row=lineIndex, column=1).value != 'AgentsEnd':
        #Definition of an agent begins
        if agentsSheet.cell(row=lineIndex, column=1).value == 'AgentBegin':
            agent = {}
            programs = []
            agent['id'] = agentsSheet.cell(row=lineIndex, column=3).value
            agent['contents'] = str(agentsSheet.cell(row=lineIndex, column=5).value).split(',')
            row = int(agentsSheet.cell(row=lineIndex, column=7).value)
            col = int(agentsSheet.cell(row=lineIndex, column=9).value)
            if row == -1:
                row = random.randrange(1, rows)
            if col == -1:
                col = random.randrange(1, columns)
            agent['coordinates'] = {
                'i': row,
                'j': col
            }
            copies = int(agentsSheet.cell(row=lineIndex, column=11).value)
        #definition of a program begins
        if agentsSheet.cell(row=lineIndex, column=2).value == 'programBegin':
            program = []
        #definition of a rule
        if agentsSheet.cell(row=lineIndex, column=2).value == 'rule':
            rule = {
                'left': '',
                'operator': agentsSheet.cell(row=lineIndex, column=4).value,
                'right':agentsSheet.cell(row=lineIndex, column=5).value
            }
            if rule['operator'] in ['u', 'd', 'l', 'r']:
                rule['left'] = str(agentsSheet.cell(row=lineIndex, column=3).value).split(',')
            else:
                rule['left'] = agentsSheet.cell(row=lineIndex, column=3).value
            program.append(rule)
        # definition of a program ends
        if agentsSheet.cell(row=lineIndex, column=2).value == 'programEnd':
            programs.append(program)
        #definition of an agent ends
        if agentsSheet.cell(row=lineIndex, column=1).value == 'AgentEnd':
            agent['programs'] = programs
            id = agent['id']
            if copies > 1:
                for i in range(copies):
                    agent['id'] = id + '_' + str(i)
                    agents.append(agent)
            else:
                agents.append(agent)
        lineIndex += 1
    print(agents)
    return agents


def getColonie(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    parameters = readParameters(wb)
    environment = readEnvironment(wb, parameters['envRows'], parameters['envColumns'], parameters['envRules'])
    agents = readAgents(wb, parameters['envRows'], parameters['envColumns'])
    return {
        'parameters': parameters,
        'environment': environment,
        'agents': agents
    }





